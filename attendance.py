from ultralytics import YOLO
import cv2
from database import *
from FacRec_LiveStream import *
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
from pytube import YouTube
import pafy
import yt_dlp
from server import *
#try using pafy for live

# # YouTube video URL
# youtube_url = 'https://youtube.com/live/lNwGmGk8acw?feature=share'

counter = 0

model = YOLO('yolov8n.pt')
people_list = []


# ydl_opts = {
#     'quiet': True,
# }

# # # Extract video stream URL using yt-dlp
# with yt_dlp.YoutubeDL(ydl_opts) as ydl:
#     info_dict = ydl.extract_info(youtube_url, download=False)
#     video_url = info_dict['url']

enter_dict = {}
leave_dict = {}

#"C:/Users/ahare/Desktop/test-4.mp4"

video_path_enter = "C:/Users/ahare/Desktop/test-4.mp4"
video_path_leave = "C:/Users/ahare/Desktop/test-4.mp4"
    
def attendance_enter():
    counter = 0
    results_generator = model.track(source=video_path_enter, classes=0, stream=True, persist=True, stream_buffer = True)
    while True:
        counter += 1
        if counter % 3 != 0:
            continue

        # Get results for the current frame
        results = next(results_generator, None)  
        if results is not None:
            # print("Iteration")
            try:
                anno_frame = results.plot()
                
                keys_to_delete = []
                for key, value in enter_dict.items():
                    if key not in results.boxes.id.tolist():
                        keys_to_delete.append(key)

                # Delete the keys outside the loop
                for key in keys_to_delete:
                    enter_dict.pop(key)

                for result in results:
                    if result.boxes.id is None:
                        continue

                    box = result.boxes.xyxy.tolist()
                    id = result.boxes.id.tolist()

                    print("This is id: ", id)
                    if enter_dict.get(id[0]) == True:
                        print("Already detected")
                        continue

                    left, top, right, bottom = map(int, box[0])

                    roi = anno_frame[top:bottom, left:right]
                    

                    if id[0] not in enter_dict:
                        enter_dict[id[0]] = False

                    detected_id = recognize_faces(roi)
                    print("Detected ids:", detected_id)
                    if detected_id:
                        enter_dict[id[0]] = True
                        process_enter_frame(detected_id[0])

                    # print(f"ids: {id} coordinates: {box}")

                # fr = cv2.resize(anno_frame, (550,550))
                # cv2.imshow("Frame", fr)
                    

                    # Draw bounding box on the frame

            except Exception as e:
                print(f"Error processing result: {e}")
                continue


        # if cv2.waitKey(1) == ord("q"):
        #     break


def process_enter_frame(detected_id):
    # Recognize faces and perform necessary processing for entering frames

    parent_id , parent_token = get_parent_id_and_token(detected_id)
    child_name , child_present = get_child_name_and_present_status(detected_id)
            
    if not child_present:
            update_child_presence_status(detected_id, True)
            entered = True
            thread = threading.Thread(
                target=create_child_attendance_excel,
                args=(detected_id, child_name, entered)
            )
            thread.start()
            expo_push_token = parent_token
            title = 'Attendance'
            body = f'You Child {child_name} Entered the bus'
            # encode_known_faces()
            send_notification(expo_push_token, title, body)

            # Print statements or additional processing
            print("Notification sent to", parent_token, "for", child_name)


def attendance_leave():
    results_generator = model.track(source=video_path_leave, classes=0, stream=True, persist=True, stream_buffer = True)
    while True:
        # Get results for the current frame
        results = next(results_generator, None)  
        if results is not None:
            try:
                anno_frame = results.plot()
                
                keys_to_delete = []
                for key, value in leave_dict.items():
                    if key not in results.boxes.id.tolist():
                        keys_to_delete.append(key)

                # Delete the keys outside the loop
                for key in keys_to_delete:
                    leave_dict.pop(key)

                for result in results:
                
                    box = result.boxes.xyxy.tolist()
                    id = result.boxes.id.tolist()
           
                    if leave_dict.get(id[0]) == True:
                        print("Already detected")
                        continue

                    left, top, right, bottom = map(int, box[0])

                    roi = anno_frame[top:bottom, left:right]
                    

                    if id[0] not in enter_dict:
                        leave_dict[id[0]] = False

                    detected_id = recognize_faces(roi)
                    print("Detected ids:", detected_id)

                    if detected_id:
                        leave_dict[id[0]] = True
                        process_leave_frame(detected_id[0])

                    # cv2.imshow("Frame", roi)
                    
                    # print(f"ids: {id} coordinates: {box}")
                    

                    # Draw bounding box on the frame

            except Exception as e:
                print(f"Error processing result: {e}")
                continue


        # if cv2.waitKey(1) == ord("q"):
        #     break


def process_leave_frame(detected_id):
    # Recognize faces and perform necessary processing for leaving frames
    parent_id , parent_token = get_parent_id_and_token(detected_id)
    child_name , child_present = get_child_name_and_present_status(detected_id)
            
    if child_present:
            update_child_presence_status(detected_id, False)
            entered = False
            thread = threading.Thread(
                target=create_child_attendance_excel,
                args=(detected_id, child_name, entered)
            )
            thread.start()
            expo_push_token = parent_token
            title = 'Attendance'
            body = f'You Child {child_name} Left the bus'
            # encode_known_faces()
            send_notification(expo_push_token, title, body)

            # Print statements or additional processing
            print("Notification sent to", parent_token, "for", child_name)


def create_child_attendance_excel(child_id, child_name, entered=False):
    # Define the path to the attendance folder
    attendance_folder = os.path.join(os.getcwd(), "attendance")

    # Check if the folder exists, if not, create it
    if not os.path.exists(attendance_folder):
        os.makedirs(attendance_folder)

    # Define the filename based on the child's ID
    excel_filename = os.path.join(attendance_folder, f"{child_id}.xlsx")

    # Check if the Excel file exists, if not, create a new one
    if not os.path.exists(excel_filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Time", "Status"])

        # Save the new workbook
        workbook.save(excel_filename)

    # Get the current date and time (you can customize this based on your needs)
    current_date = pd.Timestamp.now().date()
    current_time = pd.Timestamp.now().time()

    # Create a new row in the Excel file
    excel_data = [current_date, current_time, "Entered" if entered else "Left"]

    # Load the existing workbook
    workbook = load_workbook(excel_filename)
    sheet = workbook.active

    # Append data to the worksheet
    sheet.append(excel_data)

    # Save the updated workbook
    workbook.save(excel_filename)

    print(f"Attendance recorded for {child_name} (Child ID: {child_id})")


def send_notification(expo_push_token, title, body):
    expo_api_url = 'https://exp.host/--/api/v2/push/send'
    message = {
        'to': expo_push_token,
        'title': title,
        'body': body,
        'sound': 'default',
        'priority': 'high',
    }
    headers = {
        'Content-Type': 'application/json',
    }
    response = requests.post(expo_api_url, data=json.dumps(message), headers=headers)
    print(response.json())


# attendance_enter()



def attendance_enter(frame):
    # global counter
    # counter += 1
    # if counter % 4 != 0:
    #     return  # Skip processing for frames that are not every 4th frame

    try:
        # while True:
        #     ret, frame = cap.read()
        #     if not ret:
        #         break
        
            results = model.track(source=frame, classes=0, persist=True)  # Process the frame
            
            # anno_frame = results[0].plot()
            print("Iteration 1")

            for result in results:
                
                print("person")
                if result.boxes.id is None:
                    continue

                for people in result:
                    box = people.boxes.xyxy.tolist()
                    id = people.boxes.id.tolist()
                    print("IDS: ",id)
                    if enter_dict.get(id[0]) == True:
                        print("Already detected")
                        continue

                    left, top, right, bottom = map(int, box[0])
                    roi = frame[top:bottom, left:right]

                    if id[0] not in enter_dict:
                        enter_dict[id[0]] = False

                    detected_id = recognize_faces(roi)
                    print("Detected ids:", detected_id)
                    if detected_id:
                        enter_dict[id[0]] = True
                        process_enter_frame(detected_id[0])

            # Display the frame
            # resize = cv2.resize(anno_frame,(550,550))
            # cv2.imshow("Frame", resize)
            # cv2.waitKey(1)  # Adjust the waitKey value if needed for proper display rate

    except Exception as e:
        print(f"Error processing result: {e}")



def attendance_leave(frame):
    # global counter
    # counter += 1
    # if counter % 4 != 0:
    #     return  # Skip processing for frames that are not every 4th frame

    try:
        # while True:
        #     ret, frame = cap.read()
        #     if not ret:
        #         break
        
            results = model.track(source=frame, classes=0, persist=True)  # Process the frame
            
            # anno_frame = results[0].plot()
            print("Iteration 1")

            for result in results:
                
                print("person")
                if result.boxes.id is None:
                    continue

                for people in result:
                    box = people.boxes.xyxy.tolist()
                    id = people.boxes.id.tolist()
                    print("IDS: ",id)
                    if leave_dict.get(id[0]) == True:
                        print("Already detected")
                        continue

                    left, top, right, bottom = map(int, box[0])
                    roi = frame[top:bottom, left:right]

                    if id[0] not in enter_dict:
                        leave_dict[id[0]] = False

                    detected_id = recognize_faces(roi)
                    print("Detected ids:", detected_id)
                    if detected_id:
                        leave_dict[id[0]] = True
                        process_leave_frame(detected_id[0])

            # Display the frame
            # resize = cv2.resize(anno_frame,(550,550))
            # cv2.imshow("Frame", resize)
            # cv2.waitKey(1)  # Adjust the waitKey value if needed for proper display rate

    except Exception as e:
        print(f"Error processing result: {e}")

